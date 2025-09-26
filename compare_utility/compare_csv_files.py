import csv
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.cell import WriteOnlyCell
import os
from tempfile import NamedTemporaryFile
import logging
import time

def build_key_index_and_header(csv_path, keycol):
    """Build a mapping from keycol to file offset for a CSV file, and return header."""
    index = {}
    with open(csv_path, 'r', newline='', encoding='utf-8') as f:
        header_line = f.readline()
        header = header_line.strip().split(',')
        logging.info(f"Header for {csv_path}: {header}")
        if keycol not in header:
            logging.error(f"Key column '{keycol}' not found in {csv_path} header: {header}")
            raise ValueError(f"Key column '{keycol}' not found in {csv_path}")
        key_idx = header.index(keycol)
        while True:
            offset = f.tell()
            line = f.readline()
            if not line:
                break
            row = line.strip().split(',')
            if len(row) > key_idx:
                index[row[key_idx]] = offset
    return index, header

def get_row_dict_by_offset(csv_path, offset, header):
    with open(csv_path, 'r', newline='', encoding='utf-8') as f:
        f.seek(offset)
        line = f.readline()
        values = line.strip().split(',')
        return dict(zip(header, values))

def compare_large_csv(file1, file2, output_excel, keycol='keycol'):
    start_time = time.time()
    logging.info(f"Joining key column: '{keycol}'")
    step_start = time.time()
    idx1, header1 = build_key_index_and_header(file1, keycol)
    idx2, header2 = build_key_index_and_header(file2, keycol)
    logging.info(f"Indexing and header extraction took {time.time() - step_start:.2f} seconds.")
    all_columns = [keycol] + [col for col in header1 if col != keycol] + [col for col in header2 if col != keycol and col not in header1]
    match_columns = ['SourceFile'] + all_columns

    wb = Workbook(write_only=True)
    fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

    ws_match = wb.create_sheet('Matching Rows')
    ws_match.append(match_columns)
    ws_only1 = wb.create_sheet('Only in File1')
    ws_only1.append(match_columns)
    ws_only2 = wb.create_sheet('Only in File2')
    ws_only2.append(match_columns)

    # Matching keys
    step_start = time.time()
    logging.info(f"Comparing matching keys between files...")
    match_count = 0
    for key in idx1.keys() & idx2.keys():
        row1 = get_row_dict_by_offset(file1, idx1[key], header1)
        row2 = get_row_dict_by_offset(file2, idx2[key], header2)
        styled_row1 = [WriteOnlyCell(ws_match, value=os.path.basename(file1))]
        styled_row2 = [WriteOnlyCell(ws_match, value=os.path.basename(file2))]
        for col in all_columns:
            val1 = row1.get(col, '')
            val2 = row2.get(col, '')
            cell1 = WriteOnlyCell(ws_match, value=val1)
            cell2 = WriteOnlyCell(ws_match, value=val2)
            if col in header1 and col in header2 and val1 != val2:
                cell1.fill = fill
                cell2.fill = fill
            styled_row1.append(cell1)
            styled_row2.append(cell2)
        ws_match.append(styled_row1)
        ws_match.append(styled_row2)
        match_count += 1
        if match_count % 1000 == 0:
            logging.info(f"Processed {match_count} matching keys...")
    logging.info(f"Total matching keys processed: {match_count}")
    logging.info(f"Matching keys comparison took {time.time() - step_start:.2f} seconds.")

    # Only in file1
    step_start = time.time()
    only1_count = 0
    for key in idx1.keys() - idx2.keys():
        row1 = get_row_dict_by_offset(file1, idx1[key], header1)
        out_row = [os.path.basename(file1)] + [row1.get(col, '') for col in all_columns]
        ws_only1.append(out_row)
        only1_count += 1
        if only1_count % 1000 == 0:
            logging.info(f"Processed {only1_count} unique keys in file1...")
    logging.info(f"Total unique keys in file1: {only1_count}")
    logging.info(f"File1 unique keys processing took {time.time() - step_start:.2f} seconds.")

    # Only in file2
    step_start = time.time()
    only2_count = 0
    for key in idx2.keys() - idx1.keys():
        row2 = get_row_dict_by_offset(file2, idx2[key], header2)
        out_row = [os.path.basename(file2)] + [row2.get(col, '') for col in all_columns]
        ws_only2.append(out_row)
        only2_count += 1
        if only2_count % 1000 == 0:
            logging.info(f"Processed {only2_count} unique keys in file2...")
    logging.info(f"Total unique keys in file2: {only2_count}")
    logging.info(f"File2 unique keys processing took {time.time() - step_start:.2f} seconds.")

    # Save
    step_start = time.time()
    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']
    logging.info(f"Saving Excel output to {output_excel}")
    wb.save(output_excel)
    logging.info(f"Excel output saved successfully. Saving took {time.time() - step_start:.2f} seconds.")
    logging.info(f"Total execution time: {time.time() - start_time:.2f} seconds.")

def main():
    # User can change these values as needed for comparing other files
    file1 = 'compare_utility/file1.csv'  # Path to the first CSV file
    file2 = 'compare_utility/file2.csv'  # Path to the second CSV file
    output = 'compare_utility/comparison_output.xlsx'  # Path to the output Excel file
    keycol = 'keycol'  # Name of the key column to join on

    logging.basicConfig(level=logging.INFO, format='%(asctime)s %(levelname)s %(message)s')
    logging.info("Starting CSV comparison utility...")
    try:
        compare_large_csv(file1, file2, output, keycol)
        logging.info(f'Comparison complete. Output saved to {output}')
    except FileNotFoundError as e:
        logging.error(f"File not found: {e.filename}")
        print(f"Error: File not found: {e.filename}")
    except ValueError as e:
        logging.error(f"Value error: {e}")
        print(f"Error: {e}")
    except PermissionError as e:
        logging.error(f"Permission error: {e}")
        print(f"Error: Permission denied. {e}")
    except Exception as e:
        logging.error(f"Unexpected error: {e}", exc_info=True)
        print(f"An unexpected error occurred: {e}")

if __name__ == '__main__':
    main()
